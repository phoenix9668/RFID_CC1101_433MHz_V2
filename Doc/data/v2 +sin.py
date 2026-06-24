# 这是一个示例 Python 脚本。
import openpyxl
import numpy as np
from scipy.optimize import curve_fit



# 按 Shift+F10 执行或将其替换为您的代码。
# 按 双击 Shift 在所有地方搜索类、文件、工具窗口、操作和设置。

# 按间距中的绿色按钮以运行脚本。
if __name__ == '__main__':
    wb = openpyxl.load_workbook(filename='05_5826_0302_1549_ruminate.xlsx')
    ws = wb.active

    action_classify = {'movement': 0, 'rest': 0, 'ingestion': 0, 'climb': 0, 'other': 0, 'breathe': 0}
    threshold_judge = {'low': 0, 'normal': 0, 'abovenormal': 0, 'high': 0}
    filter_regs = [0, 0, 0]
    x_diff_regs = [212, 0]
    y_diff_regs = [212, 0]
    z_diff_regs = [212, 0]
    x_axis_info = []
    y_axis_info = []
    z_axis_info = []
    x_average_info = 0
    y_average_info = 0
    z_average_info = 0
    x_sum_info = 0
    y_sum_info = 0
    z_sum_info = 0
    x_rs_info = 0
    memory_lists = [[] for _ in range(18)]  # 将列表初始化为空列表
    txt_index = 0
    last_list = [''] * 8
    y_25 = []
    y_omega = 0
    y_rs = 0

    x_max_val = -2000  # 假设第一个元素是最大值
    x_min_val = 2000  # 假设第一个元素是最小值
    y_max_val = -2000  # 假设第一个元素是最大值
    y_min_val = 2000  # 假设第一个元素是最小值
    z_max_val = -2000  # 假设第一个元素是最大值
    z_min_val = 2000  # 假设第一个元素是最小值

    # all_average = 0
    #     all_average += y.value
    # all_average = int(all_average/ws.max_row)
    # print(all_average)
    for i in range(1, ws.max_row+1):
        # print("x value:{}".format((ws.cell(row=i, column=1)).value), end=" ")  # x value
        # print("y value:{}".format((ws.cell(row=i, column=2)).value), end=" ")  # y value
        # print("z value:{}".format((ws.cell(row=i, column=3)).value))  # z value
        x_value = (ws.cell(row=i, column=1)).value
        y_value = (ws.cell(row=i, column=2)).value
        z_value = (ws.cell(row=i, column=3)).value
        # Average Filter
        # filter_regs[2] = filter_regs[1]
        # filter_regs[1] = filter_regs[0]
        # filter_regs[0] = y.value
        # y_avg = sum(filter_regs) / 3
        x_diff_regs[1] = x_diff_regs[0]
        x_diff_regs[0] = x_value
        y_diff_regs[1] = y_diff_regs[0]
        y_diff_regs[0] = y_value
        z_diff_regs[1] = z_diff_regs[0]
        z_diff_regs[0] = z_value

        if x_value > x_max_val:
            x_max_val = x_value
        if x_value < x_min_val:
            x_min_val = x_value

        if y_value > y_max_val:
            y_max_val = y_value
        if y_value < y_min_val:
            y_min_val = y_value

        if z_value > z_max_val:
            z_max_val = z_value
        if z_value < z_min_val:
            z_min_val = z_value

        x_axis_info.append(x_diff_regs[0] - x_diff_regs[1])
        y_axis_info.append(y_diff_regs[0] - y_diff_regs[1])
        z_axis_info.append(z_diff_regs[0] - z_diff_regs[1])
        # ws.cell(row=len(y_axis_info), column=5).value = y_avg
        # print(y_axis_info)
        x_sum_info += abs(x_axis_info[-1])
        y_sum_info += abs(y_axis_info[-1])
        z_sum_info += abs(z_axis_info[-1])
        x_average_info += x_value
        y_average_info += y_value
        z_average_info += z_value
        # 将y_value添加到y_25列表中
        y_25.append(y_value)
        if len(y_25) > 25:
            y_25.pop(0)
        # print(y_average_info)

        if abs(y_axis_info[-1]) <= 10:
            threshold_judge['low'] += 1
        elif 10 < abs(y_axis_info[-1]) <= 100:
            threshold_judge['normal'] += 1
        elif 100 < abs(y_axis_info[-1]) <= 200:
            threshold_judge['abovenormal'] += 1
        elif abs(y_axis_info[-1]) > 200:
            threshold_judge['high'] += 1
        # print(threshold_judge)

        if len(y_axis_info) % 25 == 0:

            # 原始数据
            x = np.arange(1, 26)  # 使用arange生成从1到25的整数数组
            y = np.array(y_25)
            # 定义正弦函数模型
            def sine_model(x, A, omega, phi, offset):
                return A * np.sin(omega * x + phi) + offset
            initial_omega = 2 * np.pi / 10  # 假设周期大约是10（这可能需要调整）
            initial_params = [np.max(np.abs(y)), initial_omega, 0, np.mean(y)]  # [振幅, 角频率, 相位, 偏移]

            # 拟合数据
            params, covariance = curve_fit(sine_model, x, y, p0=initial_params, maxfev=10000)

            # 提取拟合参数
            A, omega, phi, offset = params

            # 将omega的值赋值给y_omega
            y_omega = omega

            # 生成拟合曲线的y值
            y_fit = sine_model(x, *params)

            # 计算R-squared（决定系数）
            y_mean = np.mean(y)
            total_ss = np.sum((y - y_mean) ** 2)
            residuals = y - y_fit
            ss_res = np.sum(residuals ** 2)
            r_squared = 1 - (ss_res / total_ss)
            y_rs = r_squared

            y_average_info = int(y_average_info / 25)
            # print(y_average_info)
            if threshold_judge['low'] >= 24:
                action_classify['rest'] += 1
                action = 1  # "rest"
            elif (threshold_judge['normal'] + threshold_judge['abovenormal']) > 11 and threshold_judge['high'] == 0 \
                    and y_average_info >= 200:
                action_classify['ingestion'] += 1
                action = 2  # "ingestion"
            elif -200 < y_average_info < 100 and (x_sum_info + y_sum_info + z_sum_info)/3 > 400 and \
                    ((x_max_val - x_min_val) + (y_max_val - y_min_val) + (z_max_val - z_min_val))/3 > 150:
                action_classify['movement'] += 1
                action = 3  # "movement"
            elif threshold_judge['high'] > 0 and y_average_info <= -200:
                action_classify['climb'] += 1
                action = 4  # "climb"
            else:
                action_classify['other'] += 1
                action = 6  # "other"
            # print(action_classify)
            # ws.cell(row=len(y_axis_info), column=5).value = action
            # 删除列表第一个位置
            memory_lists.pop(0)
            # 向循环数组添加信息
            last_list[0] = str(action)
            last_list[1] = str(y_average_info)
            last_list[2] = str(y_sum_info)
            last_list[3] = str(y_max_val - y_min_val)
            last_list[4] = str(int((x_sum_info + y_sum_info + z_sum_info)/3))
            last_list[5] = str(int(((x_max_val - x_min_val) + (y_max_val - y_min_val) + (z_max_val - z_min_val))/3))
            last_list[6] = str(y_omega)
            last_list[7] = str(y_rs)

            memory_lists.append([])
            memory_lists[-1].extend(last_list)  # 在新的空子列表的末尾添加值
            # print("begin:{}".format(memory_lists))

            if not any(isinstance(i, list) and not i for i in memory_lists):
                # print("my_list does not contain empty lists")
                # for tmpl in memory_lists[6:12]:
                #     # print("tmpl:{}".format(tmpl))
                #     if '3' in tmpl[0]:
                #         index = memory_lists.index(tmpl)
                #         if '2' in memory_lists[index - 1][0] and int(memory_lists[index][1]) > 180:
                #             memory_lists[index][0] = '2'
                movement_cnt = 0
                rest_cnt1 = 0
                rest_cnt2 = 0
                if '3' in memory_lists[9][0]:
                    for tmpl in memory_lists:
                        if '3' in tmpl[0]:
                            movement_cnt += 1
                    if movement_cnt == 1:
                        for tmpl in memory_lists[0:9]:
                            if '1' in tmpl[0]:
                                rest_cnt1 += 1
                        for tmpl in memory_lists[-9:]:
                            if '1' in tmpl[0]:
                                rest_cnt2 += 1
                        if rest_cnt1 >= 4 and rest_cnt2 >= 4:
                            memory_lists[9][0] = '1'

                movement_cnt = 0
                if '4' in memory_lists[-1][0]:
                    for tmpl in memory_lists:
                        if '3' in tmpl[0]:
                            movement_cnt += 1
                    if movement_cnt < 4:
                        memory_lists[-1][0] = '6'

                memory_lists_reversed = memory_lists[::-1]
                for tmpl in memory_lists_reversed:
                    # print("tmpl:{}".format(tmpl))
                    if '4' in tmpl[0]:
                        print("memory_list:{}".format(memory_lists))
                        print("reversed:{}".format(memory_lists_reversed))
                        index = memory_lists_reversed.index(tmpl)
                        for tmpl2 in memory_lists_reversed[index + 1:]:
                            index2 = memory_lists_reversed.index(tmpl2)
                            if '4' in tmpl2[0]:
                                memory_lists_reversed[index2][0] = '3'
                                print("reversed output:{}".format(memory_lists_reversed))
                memory_lists = memory_lists_reversed[::-1]

                ingestion_cnt = 0
                for tmpl in memory_lists:
                    if '2' in tmpl[0]:
                        ingestion_cnt += 1
                if ingestion_cnt >= 2:
                    for tmpl in memory_lists:
                        if '3' in tmpl[0]:
                            tmpl[0] = '6'

                ruminate_cnt = 0
                rest_cnt = 0
                deta_a_cnt = 0
                jicha_cnt = 0
                eighteen_average = 0
                sum_eighteen_average = 0
                sum_eighteen_y_sr = 0
                for tmpl in memory_lists:
                    if  float(tmpl[7]) > 0.9:
                        sum_eighteen_y_sr += 1
                    if '1' in tmpl[0]:
                        rest_cnt += 1
                if rest_cnt <= 4:
                    for tmpl in memory_lists:
                        if 130 < int(tmpl[4]) < 700:
                            deta_a_cnt += 1
                        if int(tmpl[5]) < 100:
                            jicha_cnt += 1
                    if deta_a_cnt >= 14 and jicha_cnt >= 14:
                        for tmpl in memory_lists:
                            eighteen_average += int(tmpl[1])
                        eighteen_average = eighteen_average / 18
                        for tmpl in memory_lists:
                            sum_eighteen_average += abs(eighteen_average - int(tmpl[1]))
                        if sum_eighteen_average <= 400 and eighteen_average < 150 and sum_eighteen_y_sr >= 8 :
                            for tmpl in memory_lists:
                                tmpl[0] = '7'
                        elif sum_eighteen_average <= 400 and eighteen_average < 150 and sum_eighteen_y_sr < 8 :
                            for tmpl in memory_lists:
                                tmpl[0] = '5'

                txt_index = 1
                ws.cell(row=len(y_axis_info) - 425, column=4).value = memory_lists[0][0]
            # print("ender:{}".format(memory_lists))

            for k in threshold_judge.keys():
                threshold_judge[k] = 0
            print(f"R-squared（决定系数）: {y_rs:.4f}, 当前行数: {i+1}")
            x_sum_info = 0
            y_sum_info = 0
            z_sum_info = 0
            x_average_info = 0
            y_average_info = 0
            z_average_info = 0
            x_max_val = -2000  # 假设第一个元素是最大值
            x_min_val = 2000  # 假设第一个元素是最小值
            y_max_val = -2000  # 假设第一个元素是最大值
            y_min_val = 2000  # 假设第一个元素是最小值
            z_max_val = -2000  # 假设第一个元素是最大值
            z_min_val = 2000  # 假设第一个元素是最小值
            y_omega = 0
            y_rs = 0
            print("***********************************")

        if txt_index == 1:
            ws.cell(row=len(y_axis_info) - 425, column=5).value = memory_lists[0][0]
        # if len(y_axis_info) > 8000:
        #     break

    wb.save(filename='05_5826_0302_1549_ruminate.xlsx')
